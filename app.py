#!/usr/bin/env python3
import os
import csv
import threading
import logging
import shutil
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_from_directory
import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook

load_dotenv()

# ----- Config -----
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
app = Flask(__name__, static_folder="static", template_folder="templates")


# --- Middleware for /CRC Prefix ---
class PrefixMiddleware(object):
    def __init__(self, app, prefix=""):
        self.app = app
        self.prefix = prefix

    def __call__(self, environ, start_response):
        if environ["PATH_INFO"].startswith(self.prefix):
            environ["PATH_INFO"] = environ["PATH_INFO"][len(self.prefix) :]
            environ["SCRIPT_NAME"] = self.prefix
            return self.app(environ, start_response)
        else:
            start_response("404", [("Content-Type", "text/plain")])
            return [b"Not Found"]


app.wsgi_app = PrefixMiddleware(app.wsgi_app, prefix="/CRC")

# --- Environment Variables ---
SNIPE_URL = os.getenv("SNIPE_URL")
SNIPE_TOKEN = os.getenv("SNIPE_API_TOKEN")
CURRENT_CSV = os.getenv("CURRENT_CSV", "current_scan.csv")
COMPLETED_FOLDER = os.getenv("COMPLETED_FOLDER", "completed_scans")
CSV_HEADERS = os.getenv(
    "CSV_HEADERS", "Equipment Type,Item Description,Serial Number,Temple Tag"
).split(",")

TEMPLATE_FILE = os.getenv("TEMPLATE_XLSX", "template.xlsx")

SNIPE_VERIFY_SSL = os.getenv("SNIPE_VERIFY_SSL", "true").lower() == "true"
SNIPE_TIMEOUT = int(os.getenv("SNIPE_TIMEOUT_SECONDS", "5"))

CSV_LOCK = threading.Lock()

if not os.path.exists(COMPLETED_FOLDER):
    os.makedirs(COMPLETED_FOLDER)


# ----- Helpers -----
def ensure_csv():
    """Ensures the CSV exists with headers."""
    with CSV_LOCK:
        if not os.path.exists(CURRENT_CSV) or os.path.getsize(CURRENT_CSV) == 0:
            with open(CURRENT_CSV, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(CSV_HEADERS)


def format_temple_tag(tag):
    """Removes 'CPH' or cuts the string off at 5 characters."""
    if not tag:
        return ""

    tag_str = str(tag).strip()

    if tag_str.upper().startswith("CPH"):
        # If it has CPH, remove the first 3 characters ("CPH")
        return tag_str[3:]
    else:
        # If it doesn't have CPH, cut it off after the 5th character
        return tag_str[:5]


def append_row(data):
    ensure_csv()
    target_serial = data.get("Serial Number", "").strip()

    with CSV_LOCK:
        # 1. READ FILE to check for duplicates (Source of Truth)
        # This fixes the multi-worker bug because they all read the same file.
        if target_serial:
            try:
                with open(CURRENT_CSV, "r", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    next(reader, None)  # Skip Header
                    for row in reader:
                        # Column 2 is Serial Number
                        if (
                            len(row) > 2
                            and row[2].strip().lower() == target_serial.lower()
                        ):
                            return False, "Duplicate Serial detected in this batch."
            except Exception as e:
                logging.error(f"Read error during dupe check: {e}")

        # 2. WRITE if safe
        try:
            row = [
                data.get("Equipment Type", ""),
                data.get("Item Description", ""),
                target_serial,
                data.get("Temple Tag", "N/A"),
            ]

            with open(CURRENT_CSV, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(row)

            return True, "Saved"
        except Exception as e:
            return False, str(e)


def lookup_snipe(term):
    if not SNIPE_URL or not SNIPE_TOKEN:
        return None
    if not term:
        return None

    base_api = SNIPE_URL.rstrip("/")
    if base_api.endswith("/hardware"):
        base_api = base_api.replace("/hardware", "")

    headers = {"Authorization": f"Bearer {SNIPE_TOKEN}", "Accept": "application/json"}

    def get_data(url, params=None):
        try:
            r = requests.get(
                url,
                headers=headers,
                params=params,
                timeout=SNIPE_TIMEOUT,
                verify=SNIPE_VERIFY_SSL,
            )
            if r.status_code == 200:
                d = r.json()
                if "id" in d:
                    return [d]
                return d.get("rows", [])
            return []
        except:
            return []

    # Try Tag -> Serial -> Search
    rows = get_data(f"{base_api}/hardware/bytag/{term}")
    if not rows:
        rows = get_data(f"{base_api}/hardware/byserial/{term}")
    if not rows:
        rows = get_data(f"{base_api}/hardware", params={"search": term, "limit": 1})

    if rows:
        hw = rows[0]
        manuf = hw.get("manufacturer", {}).get("name", "")
        model = hw.get("model", {}).get("name", "")
        full_desc = f"{manuf} {model}".strip()

        return {
            "Equipment Type": hw.get("category", {}).get("name", "Computer"),
            "Item Description": full_desc,
            "Serial Number": hw.get("serial", ""),
            "Temple Tag": format_temple_tag(hw.get("asset_tag", "")),
        }
    return None


# ----- Routes -----
@app.route("/")
def index():
    ensure_csv()
    return render_template("index.html")


@app.route("/lookup", methods=["POST"])
def api_lookup():
    data = request.json or {}
    term = data.get("serial", "").strip()

    res = lookup_snipe(term)
    if res:
        return jsonify(res)

    is_likely_tag = term.upper().startswith("CPH") or (term.isdigit() and len(term) < 8)

    return jsonify(
        {
            "Equipment Type": "Computer",
            "Item Description": "",
            "Serial Number": ("" if is_likely_tag else term),
            "Temple Tag": format_temple_tag(
                term if is_likely_tag else ""
            ),
            "found_in_snipe": False,
        }
    )


@app.route("/add", methods=["POST"])
def api_add():
    data = request.json or {}
    success, msg = append_row(data)
    return jsonify({"ok": success, "error": msg})


@app.route("/recent", methods=["GET"])
def api_recent():
    ensure_csv()
    try:
        with CSV_LOCK:
            with open(CURRENT_CSV, "r", encoding="utf-8") as f:
                reader = list(csv.reader(f))
                if len(reader) > 1:
                    last_rows = reader[1:][-5:]
                    last_rows.reverse()
                    return jsonify({"items": last_rows})
    except:
        pass
    return jsonify({"items": []})


@app.route("/finalize", methods=["POST"])
def api_finalize():
    with CSV_LOCK:
        if not os.path.exists(CURRENT_CSV):
            return jsonify({"error": "No data to finalize"}), 400

        today_str = datetime.now().strftime("%Y%m%d")
        base_name = f"{today_str}-cph-crc"
        filename = f"{base_name}.xlsx"
        dest_path = os.path.join(COMPLETED_FOLDER, filename)

        counter = 1
        while os.path.exists(dest_path):
            filename = f"{base_name}-{counter}.xlsx"
            dest_path = os.path.join(COMPLETED_FOLDER, filename)
            counter += 1

        try:
            # 1. Load the template if it exists, otherwise fallback to a blank workbook
            if os.path.exists(TEMPLATE_FILE):
                wb = load_workbook(TEMPLATE_FILE)
                ws = wb.active
            else:
                logging.warning(
                    f"Template {TEMPLATE_FILE} not found. Creating a blank one."
                )
                wb = Workbook()
                ws = wb.active
                ws.append(CSV_HEADERS)  # Write headers manually if no template

            # 2. Read CSV and append to worksheet
            with open(CURRENT_CSV, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader, None)  # Skip the CSV header row

                # Append each data row to the template
                for row in reader:
                    ws.append(row)

            # 3. Save as the new destination file
            wb.save(dest_path)

            # 4. Wipe the current CSV to start fresh
            os.remove(CURRENT_CSV)

            return jsonify({"ok": True, "filename": filename})

        except Exception as e:
            logging.error(f"Finalize Error: {e}")
            return jsonify({"error": str(e)}), 500


@app.route("/reset_batch", methods=["POST"])
def api_reset_batch():
    """Wipes the current CSV."""
    with CSV_LOCK:
        with open(CURRENT_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(CSV_HEADERS)
        # No memory to clear

    return jsonify({"ok": True})


@app.route("/completed_files", methods=["GET"])
def api_completed_files():
    files = []
    if os.path.exists(COMPLETED_FOLDER):
        files = [
            f
            for f in os.listdir(COMPLETED_FOLDER)
            if f.endswith(".xlsx") or f.endswith(".csv")
        ]
    files.sort(reverse=True)
    return jsonify({"files": files})


@app.route("/download/<path:filename>")
def download_file(filename):
    return send_from_directory(COMPLETED_FOLDER, filename, as_attachment=True)


@app.route("/delete_row", methods=["POST"])
def api_delete_row():
    data = request.json or {}
    serial_to_delete = data.get("serial", "").strip().lower()

    with CSV_LOCK:
        if not os.path.exists(CURRENT_CSV):
            return jsonify({"error": "No active batch"}), 400

        try:
            with open(CURRENT_CSV, "r", encoding="utf-8") as f:
                rows = list(csv.reader(f))

            if not rows:
                return jsonify({"ok": True})

            header = rows[0]
            # Keep the header, and keep all rows where the serial (column index 2) DOES NOT match
            new_rows = [header] + [
                r
                for r in rows[1:]
                if len(r) > 2 and r[2].strip().lower() != serial_to_delete
            ]

            with open(CURRENT_CSV, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerows(new_rows)

            return jsonify({"ok": True})
        except Exception as e:
            return jsonify({"error": str(e)}), 500


@app.route("/edit_row", methods=["POST"])
def api_edit_row():
    data = request.json or {}
    old_serial = data.get("old_serial", "").strip().lower()
    new_data = data.get("new_data", {})

    with CSV_LOCK:
        try:
            with open(CURRENT_CSV, "r", encoding="utf-8") as f:
                rows = list(csv.reader(f))

            if not rows:
                return jsonify({"error": "No active batch"}), 400

            header = rows[0]
            new_rows = [header]

            # Loop through and replace the row that matches the old serial
            for r in rows[1:]:
                if len(r) > 2 and r[2].strip().lower() == old_serial:
                    new_rows.append(
                        [
                            new_data.get("Equipment Type", r[0]),
                            new_data.get("Item Description", r[1]),
                            new_data.get(
                                "Serial Number", r[2]
                            ),  # Assuming they might change the serial too
                            new_data.get("Temple Tag", len(r) > 3 and r[3] or "N/A"),
                        ]
                    )
                else:
                    new_rows.append(r)

            with open(CURRENT_CSV, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerows(new_rows)

            return jsonify({"ok": True})
        except Exception as e:
            return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    ensure_csv()
    app.run(host="0.0.0.0", port=5000, debug=True)
